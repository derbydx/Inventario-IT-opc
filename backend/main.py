from fastapi import FastAPI, Depends, HTTPException, status, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from sqlalchemy.orm import Session
from typing import List
from datetime import datetime, date
import traceback, os

import models
import schemas
from database import engine, get_db
from auth import hash_password, verify_password, create_access_token, get_current_admin, require_permission

# Inicializar la base de datos SQLite
models.Base.metadata.create_all(bind=engine)

# ==========================================
# SEEDER AUTOMÁTICO: CREAR PRIMER ADMIN SI NO EXISTE
# ==========================================
# ==========================================
# SEEDER AUTOMÁTICO REFORZADO (REPARA CONTRASEÑAS)
# ==========================================
def seed_groups_and_admin():
    db = get_db().__next__()
    try:
        # Migrar esquema: agregar columnas group_id e is_active a admins si no existen
        from sqlalchemy import inspect, text
        inspector = inspect(db.bind)
        admins_cols = [c["name"] for c in inspector.get_columns("admins")]
        if "group_id" not in admins_cols:
            db.execute(text("ALTER TABLE admins ADD COLUMN group_id INTEGER REFERENCES groups(id)"))
        if "is_active" not in admins_cols:
            db.execute(text("ALTER TABLE admins ADD COLUMN is_active BOOLEAN DEFAULT 1"))

        assets_cols = [c["name"] for c in inspector.get_columns("assets")]
        if "repair_reason" not in assets_cols:
            db.execute(text("ALTER TABLE assets ADD COLUMN repair_reason TEXT"))
        if "repair_left_by_id" not in assets_cols:
            db.execute(text("ALTER TABLE assets ADD COLUMN repair_left_by_id INTEGER REFERENCES persons(id)"))
        if "repair_technician_id" not in assets_cols:
            db.execute(text("ALTER TABLE assets ADD COLUMN repair_technician_id INTEGER REFERENCES admins(id)"))
        db.commit()
        
        # Crear grupos por defecto si no existen
        default_groups = [
            {"name": "Administrador", "description": "Acceso total al sistema", "can_view": True, "can_create": True, "can_edit": True, "can_delete": True, "can_checkout": True, "can_import_export": True, "can_manage_users": True, "is_default": False},
            {"name": "Nivel 2", "description": "Puede crear, editar y eliminar activos, hacer transacciones", "can_view": True, "can_create": True, "can_edit": True, "can_delete": True, "can_checkout": True, "can_import_export": False, "can_manage_users": False, "is_default": False},
            {"name": "Tecnico", "description": "Puede registrar y editar activos, hacer transacciones", "can_view": True, "can_create": True, "can_edit": True, "can_delete": False, "can_checkout": True, "can_import_export": False, "can_manage_users": False, "is_default": False},
            {"name": "Almacen", "description": "Solo checkout/checkin de activos existentes", "can_view": True, "can_create": False, "can_edit": False, "can_delete": False, "can_checkout": True, "can_import_export": False, "can_manage_users": False, "is_default": False},
            {"name": "Solo Lectura", "description": "Puede ver el sistema sin realizar cambios", "can_view": True, "can_create": False, "can_edit": False, "can_delete": False, "can_checkout": False, "can_import_export": False, "can_manage_users": False, "is_default": True},
        ]
        for g in default_groups:
            existing = db.query(models.Group).filter(models.Group.name == g["name"]).first()
            if not existing:
                db.add(models.Group(**g))
        db.commit()
        
        # Asignar grupo Administrador a derby_admin
        admin = db.query(models.Admin).filter(models.Admin.username == "derby_admin").first()
        admin_group = db.query(models.Group).filter(models.Group.name == "Administrador").first()
        
        if not admin:
            print("Sembrando cuenta de administrador inicial...")
            nuevo_admin = models.Admin(
                username="derby_admin",
                email="derby@empresa.com",
                password_hash=hash_password("admin123"),
                role="Administrator",
                group_id=admin_group.id if admin_group else None,
                is_active=True
            )
            db.add(nuevo_admin)
            db.commit()
            print("Usuario 'derby_admin' creado con exito. Contrasena: admin123")
        else:
            admin.password_hash = hash_password("admin123")
            if not admin.group_id and admin_group:
                admin.group_id = admin_group.id
            if admin.is_active is None:
                admin.is_active = True
            db.commit()
            print("Usuario 'derby_admin' detectado. Contrasena restablecida en: admin123")

        # Sembrar categorias desde valores existentes en assets
        from sqlalchemy import text as sa_text
        inspector = inspect(db.bind)
        if "categories" not in [c["name"] for c in inspector.get_columns("categories")]:
            pass  # tabla aun no creada (se crea con models)
        existing = db.query(models.Category).count()
        if existing == 0:
            distinct_vals = db.execute(sa_text("SELECT DISTINCT category FROM assets WHERE category IS NOT NULL AND category != ''")).fetchall()
            for (cat_name,) in distinct_vals:
                db.add(models.Category(name=cat_name))
            db.commit()
            print(f"Sembradas {len(distinct_vals)} categorias desde assets existentes.")
            
    except Exception as e:
        print(f"No se pudo ejecutar el seed: {e}")
        import traceback
        traceback.print_exc()
    finally:
        db.close()

# Ejecutamos el Seeder al arrancar el backend
seed_groups_and_admin()

app = FastAPI(
    title="IT Asset Manager API",
    description="Backend completo estilo AssetTiger",
    version="1.0.0"
)

# ==========================================
# CONFIGURACIÓN DE CORS BLINDADA
# ==========================================
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==========================================
# CAPTURADOR GLOBAL DE ERRORES
# ==========================================
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    print("ERROR CRITICO DETECTADO EN EL BACKEND:")
    traceback.print_exc()
    return JSONResponse(
        status_code=500,
        content={"detail": f"Error interno en Python/SQLite: {str(exc)}"},
    )

# ==========================================
# HEALTH CHECK
# ==========================================
@app.get("/api/health", tags=["Diagnóstico"])
def health_check():
    return {"status": "ok"}

# ==========================================
# STATIC FILES - catch-all for frontend SPA
# ==========================================
from starlette.middleware.base import BaseHTTPMiddleware

FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "static")

class SPAStaticFiles(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        if response.status_code == 404 and request.method in ("GET", "HEAD"):
            file_path = os.path.join(FRONTEND_DIR, request.url.path.lstrip("/"))
            if os.path.isfile(file_path):
                return FileResponse(file_path)
            index_path = os.path.join(FRONTEND_DIR, "index.html")
            if os.path.isfile(index_path):
                return FileResponse(index_path)
        return response

app.add_middleware(SPAStaticFiles)

# ==========================================
# 1. ENDPOINTS: SITIOS (SITES)
# ==========================================
@app.post("/sites/", response_model=schemas.SiteResponse, status_code=status.HTTP_201_CREATED, tags=["Sitios y Ubicaciones"])
def create_site(site: schemas.SiteBase, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_create"))):
    db_site = db.query(models.Site).filter(models.Site.site_name == site.site_name).first()
    if db_site:
        raise HTTPException(status_code=400, detail="El sitio ya existe")
    nuevo_sitio = models.Site(**site.model_dump())
    db.add(nuevo_sitio)
    db.commit()
    db.refresh(nuevo_sitio)
    return nuevo_sitio

@app.get("/sites/", response_model=List[schemas.SiteResponse], tags=["Sitios y Ubicaciones"])
def list_sites(db: Session = Depends(get_db)):
    return db.query(models.Site).all()

@app.put("/sites/{site_id}", response_model=schemas.SiteResponse, tags=["Sitios y Ubicaciones"])
def update_site(site_id: int, site: schemas.SiteBase, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_edit"))):
    db_site = db.query(models.Site).filter(models.Site.id == site_id).first()
    if not db_site: raise HTTPException(404, "Sitio no encontrado")
    for k, v in site.model_dump().items():
        setattr(db_site, k, v)
    db.commit(); db.refresh(db_site)
    return db_site

# ==========================================
# 2. ENDPOINTS: DEPARTAMENTOS
# ==========================================
@app.post("/departments/", response_model=schemas.DepartmentResponse, status_code=status.HTTP_201_CREATED, tags=["Catálogos"])
def create_department(dept: schemas.DepartmentBase, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_create"))):
    db_dept = db.query(models.Department).filter(models.Department.department_name == dept.department_name).first()
    if db_dept:
        raise HTTPException(status_code=400, detail="El departamento ya existe")
    nuevo_depto = models.Department(**dept.model_dump())
    db.add(nuevo_depto)
    db.commit()
    db.refresh(nuevo_depto)
    return nuevo_depto

@app.get("/departments/", response_model=List[schemas.DepartmentResponse], tags=["Catálogos"])
def list_departments(db: Session = Depends(get_db)):
    return db.query(models.Department).all()

@app.put("/departments/{dept_id}", response_model=schemas.DepartmentResponse, tags=["Catálogos"])
def update_department(dept_id: int, dept: schemas.DepartmentBase, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_edit"))):
    db_dept = db.query(models.Department).filter(models.Department.id == dept_id).first()
    if not db_dept: raise HTTPException(404, "Departamento no encontrado")
    for k, v in dept.model_dump().items():
        setattr(db_dept, k, v)
    db.commit(); db.refresh(db_dept)
    return db_dept

# ==========================================
# 4. ENDPOINTS: CATEGORÍAS (desde distinct de assets)
# ==========================================
@app.get("/categories/distinct/", tags=["Catálogos"])
def list_categories_distinct(db: Session = Depends(get_db)):
    results = db.query(models.Asset.category).distinct(models.Asset.category).order_by(models.Asset.category).all()
    return [r[0] for r in results if r[0]]

@app.get("/categories/", response_model=List[schemas.CategoryResponse], tags=["Catálogos"])
def list_categories(db: Session = Depends(get_db)):
    return db.query(models.Category).order_by(models.Category.name).all()

@app.post("/categories/", response_model=schemas.CategoryResponse, status_code=201, tags=["Catálogos"])
def create_category(cat: schemas.CategoryCreate, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_edit"))):
    name = cat.name.strip()
    if not name:
        raise HTTPException(400, "El nombre de la categoria no puede estar vacio")
    existing = db.query(models.Category).filter(models.Category.name == name).first()
    if existing:
        raise HTTPException(400, "Ya existe una categoria con ese nombre")
    db_cat = models.Category(name=name)
    db.add(db_cat)
    db.commit()
    db.refresh(db_cat)
    return db_cat

@app.put("/categories/{cat_id}", response_model=schemas.CategoryResponse, tags=["Catálogos"])
def update_category(cat_id: int, cat: schemas.CategoryCreate, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_edit"))):
    db_cat = db.query(models.Category).filter(models.Category.id == cat_id).first()
    if not db_cat:
        raise HTTPException(404, "Categoria no encontrada")
    name = cat.name.strip()
    if not name:
        raise HTTPException(400, "El nombre no puede estar vacio")
    conflict = db.query(models.Category).filter(models.Category.name == name, models.Category.id != cat_id).first()
    if conflict:
        raise HTTPException(400, "Ya existe otra categoria con ese nombre")
    db_cat.name = name
    db.commit()
    db.refresh(db_cat)
    return db_cat

@app.delete("/categories/{cat_id}", tags=["Catálogos"])
def delete_category(cat_id: int, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_delete"))):
    db_cat = db.query(models.Category).filter(models.Category.id == cat_id).first()
    if not db_cat:
        raise HTTPException(404, "Categoria no encontrada")
    in_use = db.query(models.Asset).filter(models.Asset.category == db_cat.name).first()
    if in_use:
        raise HTTPException(400, f"No se puede eliminar la categoria '{db_cat.name}' porque hay activos que la usan")
    db.delete(db_cat)
    db.commit()
    return {"detail": "Categoria eliminada"}

# ==========================================
# 5. ENDPOINTS: ADMINISTRADORES (ADMINS)
# ==========================================
@app.get("/admins/", response_model=List[schemas.AdminResponse], tags=["Seguridad y Administradores"])
def list_admins(db: Session = Depends(get_db)):
    return db.query(models.Admin).all()

# ==========================================
# 6. ENDPOINTS: EMPLEADOS / PERSONAS (PERSONS)
# ==========================================
@app.post("/persons/", response_model=schemas.PersonResponse, status_code=status.HTTP_201_CREATED, tags=["Directorio de Personal"])
def create_person(person: schemas.PersonCreate, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_create"))):
    db_person = db.query(models.Person).filter((models.Person.email == person.email) | (models.Person.employee_id == person.employee_id)).first()
    if db_person:
        raise HTTPException(status_code=400, detail="El Employee ID o Correo ya existen en el directorio")
    
    nueva_persona = models.Person(**person.model_dump())
    db.add(nueva_persona)
    db.commit()
    db.refresh(nueva_persona)
    return nueva_persona

@app.get("/persons/", response_model=List[schemas.PersonResponse], tags=["Directorio de Personal"])
def list_persons(db: Session = Depends(get_db)):
    return db.query(models.Person).all()

@app.put("/persons/{person_id}", response_model=schemas.PersonResponse, tags=["Directorio de Personal"])
def update_person(person_id: int, person: schemas.PersonUpdate, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_edit"))):
    db_person = db.query(models.Person).filter(models.Person.id == person_id).first()
    if not db_person: raise HTTPException(404, "Empleado no encontrado")
    if person.email:
        existing = db.query(models.Person).filter(models.Person.email == person.email, models.Person.id != person_id).first()
        if existing: raise HTTPException(400, "El email ya esta en uso por otro empleado")
    if person.employee_id:
        existing = db.query(models.Person).filter(models.Person.employee_id == person.employee_id, models.Person.id != person_id).first()
        if existing: raise HTTPException(400, "El Employee ID ya esta en uso por otro empleado")
    for k, v in person.model_dump(exclude_unset=True).items():
        setattr(db_person, k, v)
    db.commit(); db.refresh(db_person)
    return db_person

# ==========================================
# 7. ENDPOINTS: ACTIVOS (ASSETS)
# ==========================================
@app.post("/assets/", response_model=schemas.AssetResponse, status_code=status.HTTP_201_CREATED, tags=["Gestión de Activos"])
def create_asset(asset: schemas.AssetCreate, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_create"))):
    db_asset = db.query(models.Asset).filter(models.Asset.asset_tag_id == asset.asset_tag_id).first()
    if db_asset:
        raise HTTPException(status_code=400, detail="El Asset Tag ID ya está registrado")
    
    nuevo_activo = models.Asset(**asset.model_dump())
    db.add(nuevo_activo)
    db.commit()
    db.refresh(nuevo_activo)
    return nuevo_activo

@app.get("/assets/", response_model=List[schemas.AssetResponse], tags=["Gestión de Activos"])
def list_assets(
    search: str = None,
    search_condition: str = "contains",
    search_field: str = None,
    status: str = None,
    category: str = None,
    site_id: int = None,
    department_id: int = None,
    person_id: int = None,
    purchased_from: str = None,
    date_field: str = "purchase_date",
    date_from: str = None,
    date_to: str = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    query = db.query(models.Asset)
    if search and search_field:
        column = getattr(models.Asset, search_field, None)
        if column:
            if search_condition == "exact":
                query = query.filter(column == search)
            elif search_condition == "startswith":
                query = query.filter(column.like(f"{search}%"))
            elif search_condition == "endswith":
                query = query.filter(column.like(f"%{search}"))
            else:
                query = query.filter(column.like(f"%{search}%"))
    elif search:
        like = f"%{search}%"
        query = query.filter(
            models.Asset.asset_tag_id.like(like) |
            models.Asset.asset_description.like(like) |
            models.Asset.brand.like(like) |
            models.Asset.model.like(like) |
            models.Asset.serial_no.like(like) |
            models.Asset.category.like(like)
        )
    if status:
        query = query.filter(models.Asset.status == status)
    if category:
        query = query.filter(models.Asset.category == category)
    if site_id:
        query = query.filter(models.Asset.site_id == site_id)
    if department_id:
        query = query.join(models.Person, models.Asset.person_id == models.Person.id).filter(models.Person.department_id == department_id)
    if person_id:
        query = query.filter(models.Asset.person_id == person_id)
    if purchased_from:
        query = query.filter(models.Asset.purchased_from.like(f"%{purchased_from}%"))
    if date_from and date_field == "purchase_date":
        query = query.filter(models.Asset.purchase_date >= datetime.strptime(date_from, "%Y-%m-%d").date())
    if date_to and date_field == "purchase_date":
        query = query.filter(models.Asset.purchase_date <= datetime.strptime(date_to, "%Y-%m-%d").date())
    if date_from and date_field == "assigned_date":
        subq = db.query(models.History.asset_id).filter(
            models.History.tipo_accion == "Checkout",
            models.History.fecha_accion >= date_from
        ).subquery()
        query = query.filter(models.Asset.id.in_(subq))
    if date_to and date_field == "assigned_date":
        subq = db.query(models.History.asset_id).filter(
            models.History.tipo_accion == "Checkout",
            models.History.fecha_accion <= date_to
        ).subquery()
        query = query.filter(models.Asset.id.in_(subq))
    return query.offset(skip).limit(limit).all()

@app.get("/assets/count/", tags=["Gestión de Activos"])
def count_assets(
    search: str = None,
    search_condition: str = "contains",
    search_field: str = None,
    status: str = None,
    category: str = None,
    site_id: int = None,
    department_id: int = None,
    person_id: int = None,
    purchased_from: str = None,
    date_field: str = "purchase_date",
    date_from: str = None,
    date_to: str = None,
    db: Session = Depends(get_db)
):
    query = db.query(models.Asset)
    if search and search_field:
        column = getattr(models.Asset, search_field, None)
        if column:
            if search_condition == "exact":
                query = query.filter(column == search)
            elif search_condition == "startswith":
                query = query.filter(column.like(f"{search}%"))
            elif search_condition == "endswith":
                query = query.filter(column.like(f"%{search}"))
            else:
                query = query.filter(column.like(f"%{search}%"))
    elif search:
        like = f"%{search}%"
        query = query.filter(
            models.Asset.asset_tag_id.like(like) |
            models.Asset.asset_description.like(like) |
            models.Asset.brand.like(like) |
            models.Asset.model.like(like) |
            models.Asset.serial_no.like(like) |
            models.Asset.category.like(like)
        )
    if status:
        query = query.filter(models.Asset.status == status)
    if category:
        query = query.filter(models.Asset.category == category)
    if site_id:
        query = query.filter(models.Asset.site_id == site_id)
    if department_id:
        query = query.join(models.Person, models.Asset.person_id == models.Person.id).filter(models.Person.department_id == department_id)
    if person_id:
        query = query.filter(models.Asset.person_id == person_id)
    if purchased_from:
        query = query.filter(models.Asset.purchased_from.like(f"%{purchased_from}%"))
    if date_from and date_field == "purchase_date":
        query = query.filter(models.Asset.purchase_date >= datetime.strptime(date_from, "%Y-%m-%d").date())
    if date_to and date_field == "purchase_date":
        query = query.filter(models.Asset.purchase_date <= datetime.strptime(date_to, "%Y-%m-%d").date())
    if date_from and date_field == "assigned_date":
        subq = db.query(models.History.asset_id).filter(
            models.History.tipo_accion == "Checkout",
            models.History.fecha_accion >= date_from
        ).subquery()
        query = query.filter(models.Asset.id.in_(subq))
    if date_to and date_field == "assigned_date":
        subq = db.query(models.History.asset_id).filter(
            models.History.tipo_accion == "Checkout",
            models.History.fecha_accion <= date_to
        ).subquery()
        query = query.filter(models.Asset.id.in_(subq))
    return {"count": query.count()}

@app.get("/assets/{asset_id}", tags=["Assets"])
def get_asset(asset_id: int, db: Session = Depends(get_db)):
    asset = db.query(models.Asset).filter(models.Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(404, "Asset no encontrado")
    return asset

# ==========================================
# 8. ACCIONES DE INVENTARIO: CHECKOUT Y CHECKIN
# ==========================================
@app.post("/assets/{asset_id}/checkout", tags=["Acciones de Inventario"])
def asset_checkout(asset_id: int, person_id: int, notas: str = None, db: Session = Depends(get_db), current_admin: models.Admin = Depends(require_permission("can_checkout"))):
    asset = db.query(models.Asset).filter(models.Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Activo no encontrado")
    
    if asset.status == "Checkout":
        raise HTTPException(status_code=400, detail="El activo ya se encuentra asignado (Checkout)")
        
    person = db.query(models.Person).filter(models.Person.id == person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail="Empleado no encontrado")

    estado_anterior = asset.status
    asset.status = "Checkout"
    asset.person_id = person_id
    
    registro_historial = models.History(
        asset_id=asset.id,
        asignado_a_id=person_id,
        realizado_por_id=current_admin.id,
        tipo_accion="Checkout",
        estado_anterior=estado_anterior,
        estado_nuevo="Checkout",
        notas_detalle=notas or f"Equipo asignado a {person.full_name}"
    )
    
    db.add(registro_historial)
    db.commit()
    db.refresh(asset)
    return {"message": f"Asset {asset.asset_tag_id} asignado exitosamente", "asset_status": asset.status}

@app.post("/assets/{asset_id}/checkin", tags=["Acciones de Inventario"])
def asset_checkin(asset_id: int, nuevo_estado: str = "Available", notas: str = None, db: Session = Depends(get_db), current_admin: models.Admin = Depends(require_permission("can_checkout"))):
    asset = db.query(models.Asset).filter(models.Asset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Activo no encontrado")

    estado_anterior = asset.status
    persona_que_devuelve = asset.person_id

    asset.status = nuevo_estado
    asset.person_id = None 
    
    registro_historial = models.History(
        asset_id=asset.id,
        asignado_a_id=persona_que_devuelve,
        realizado_por_id=current_admin.id,
        tipo_accion="Check in",
        estado_anterior=estado_anterior,
        estado_nuevo=nuevo_estado,
        notas_detalle=notas or f"Equipo recibido en almacén con estado: {nuevo_estado}"
    )
    
    db.add(registro_historial)
    db.commit()
    db.refresh(asset)
    return {"message": f"Asset {asset.asset_tag_id} recibido exitosamente", "asset_status": asset.status}

@app.get("/history/", response_model=List[schemas.HistoryResponse], tags=["Acciones de Inventario"])
def view_history(
    skip: int = 0,
    limit: int = 200,
    db: Session = Depends(get_db)
):
    return db.query(models.History).order_by(models.History.id.desc()).offset(skip).limit(limit).all()

@app.get("/history/count/", tags=["Acciones de Inventario"])
def count_history(db: Session = Depends(get_db)):
    return {"count": db.query(models.History).count()}

# ==========================================
# 9. EDICIÓN Y ELIMINACIÓN (SOFT DELETE)
# ==========================================
@app.put("/assets/{asset_id}", response_model=schemas.AssetResponse, tags=["Gestión de Activos"])
def update_asset(asset_id: int, asset_update: schemas.AssetCreate, db: Session = Depends(get_db), current_admin: models.Admin = Depends(require_permission("can_edit"))):
    db_asset = db.query(models.Asset).filter(models.Asset.id == asset_id).first()
    if not db_asset:
        raise HTTPException(status_code=404, detail="Activo no encontrado")
    
    etiquetas = {
        "asset_tag_id": "Asset Tag",
        "asset_description": "Descripción",
        "brand": "Marca",
        "model": "Modelo",
        "serial_no": "Número de Serie",
        "category": "Categoria",
        "site_id": "ID de Sitio",
        "notas_adicionales": "Notas",
        "numero_telefono": "Teléfono"
    }
    
    lista_cambios = []
    datos_nuevos = asset_update.model_dump()
    
    for campo, nuevo_valor in datos_nuevos.items():
        viejo_valor = getattr(db_asset, campo, None)
        if viejo_valor != nuevo_valor:
            nombre_limpio = etiquetas.get(campo, campo)
            lista_cambios.append(f"{nombre_limpio} cambiado de '{viejo_valor}' a '{nuevo_valor}'")
            
    repair_statuses = ("Under repair", "GarantiaSD")

    if asset_update.status in repair_statuses:
        db_asset.repair_reason = asset_update.repair_reason
        db_asset.repair_left_by_id = asset_update.repair_left_by_id
        db_asset.repair_technician_id = asset_update.repair_technician_id
    elif db_asset.status not in repair_statuses and asset_update.status not in repair_statuses:
        pass
    else:
        db_asset.repair_reason = None
        db_asset.repair_left_by_id = None
        db_asset.repair_technician_id = None

    for key, value in datos_nuevos.items():
        if key in ("repair_reason", "repair_left_by_id", "repair_technician_id"):
            continue
        setattr(db_asset, key, value)

    if db_asset.status not in ("Checkout", "Reserved"):
        db_asset.person_id = None

    if lista_cambios:
        nota_auditoria = "Edicion de propiedades: " + " | ".join(lista_cambios)
    else:
        nota_auditoria = "Formulario de edicion guardado sin cambios en los valores."
    
    tipo_accion = "Modified"
    if db_asset.status in repair_statuses:
        tipo_accion = db_asset.status
    
    registro_historial = models.History(
        asset_id=db_asset.id,
        asignado_a_id=db_asset.person_id,
        realizado_por_id=current_admin.id,
        tipo_accion=tipo_accion,
        estado_anterior=db_asset.status,
        estado_nuevo=db_asset.status,
        notas_detalle=nota_auditoria
    )
    
    db.add(registro_historial)
    db.commit()
    db.refresh(db_asset)
    return db_asset

@app.delete("/assets/{asset_id}", tags=["Gestión de Activos"])
def delete_asset(asset_id: int, db: Session = Depends(get_db), current_admin: models.Admin = Depends(require_permission("can_delete"))):
    db_asset = db.query(models.Asset).filter(models.Asset.id == asset_id).first()
    if not db_asset:
        raise HTTPException(status_code=404, detail="Activo no encontrado")
    
    estado_anterior = db_asset.status
    persona_que_tenia = db_asset.person_id

    db_asset.status = "Archived"
    db_asset.person_id = None 

    registro_historial = models.History(
        asset_id=db_asset.id,
        asignado_a_id=persona_que_tenia,
        realizado_por_id=current_admin.id,
        tipo_accion="Archived",
        estado_anterior=estado_anterior,
        estado_nuevo="Archived",
        notas_detalle="Activo enviado a Eliminados Recientemente (Baja de Inventario)"
    )
    
    db.add(registro_historial)
    db.commit()
    return {"message": "Activo movido a eliminados recientemente"}

# ==========================================
# 11. IMPORTACION Y EXPORTACION EXCEL
# ==========================================
from fastapi import UploadFile, File
from fastapi.responses import StreamingResponse, Response
import io, openpyxl
from openpyxl import Workbook

def _make_excel(headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ---------- EXPORTAR ----------
@app.get("/export/assets/", tags=["Import/Export"])
def export_assets(db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_import_export"))):
    q = db.query(models.Asset).all()
    rows = []
    for a in q:
        p = db.query(models.Person).filter(models.Person.id == a.person_id).first() if a.person_id else None
        s = db.query(models.Site).filter(models.Site.id == a.site_id).first() if a.site_id else None
        rows.append((a.asset_tag_id, a.asset_description, a.brand, a.model, a.serial_no, a.category or "", s.site_name if s else "", p.email if p else ""))
    buf = _make_excel(["AssetTag", "Descripcion", "Marca", "Modelo", "Serie", "Categoria", "Sitio", "AsignadoA"], rows)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=activos.xlsx"})

@app.get("/export/persons/", tags=["Import/Export"])
def export_persons(db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_import_export"))):
    q = db.query(models.Person).all()
    rows = []
    for p in q:
        dept = db.query(models.Department).filter(models.Department.id == p.department_id).first()
        site = db.query(models.Site).filter(models.Site.id == p.site_id).first()
        rows.append((p.full_name, p.email, p.employee_id, p.title or "", p.phone or "", dept.department_name if dept else "", site.site_name if site else ""))
    buf = _make_excel(["Nombre", "Email", "EmployeeID", "Titulo", "Telefono", "Departamento", "Sitio"], rows)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=empleados.xlsx"})

@app.get("/export/sites/", tags=["Import/Export"])
def export_sites(db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_import_export"))):
    q = db.query(models.Site).all()
    rows = [(s.site_name, s.city or "", s.country or "") for s in q]
    buf = _make_excel(["Sitio", "Ciudad", "Pais"], rows)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=sitios.xlsx"})

# ---------- PLANTILLAS ----------
@app.get("/export/assets/template/", tags=["Import/Export"])
def template_assets(): return StreamingResponse(_make_excel(["AssetTag", "Descripcion", "Marca", "Modelo", "Serie", "Categoria", "Sitio", "AsignadoA"], []), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=plantilla_activos.xlsx"})

@app.get("/export/persons/template/", tags=["Import/Export"])
def template_persons(): return StreamingResponse(_make_excel(["Nombre", "Email", "EmployeeID", "Titulo", "Telefono", "Departamento", "Sitio"], []), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=plantilla_empleados.xlsx"})

@app.get("/export/sites/template/", tags=["Import/Export"])
def template_sites(): return StreamingResponse(_make_excel(["Sitio", "Ciudad", "Pais"], []), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=plantilla_sitios.xlsx"})

# ---------- IMPORTAR ----------
def _cell(val): return str(val).strip() if val is not None else ""

@app.post("/import/sites/", tags=["Import/Export"])
def import_sites(file: UploadFile = File(...), db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_import_export"))):
    wb = openpyxl.load_workbook(file.file)
    ws = wb.active
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for col in ["Sitio"]:
        if col not in h: raise HTTPException(400, f"Falta columna '{col}'")
    i_sitio, i_ciudad, i_pais = h.index("Sitio"), h.index("Ciudad") if "Ciudad" in h else None, h.index("Pais") if "Pais" in h else None
    ok = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row): continue
        name = _cell(row[i_sitio])
        if not name: raise HTTPException(400, f"Fila {ok+2}: Sitio vacio")
        if db.query(models.Site).filter(models.Site.site_name == name).first(): raise HTTPException(400, f"Fila {ok+2}: El sitio '{name}' ya existe")
        db.add(models.Site(site_name=name, city=_cell(row[i_ciudad]) if i_ciudad is not None else None, country=_cell(row[i_pais]) if i_pais is not None else None))
        ok += 1
    db.commit()
    return {"importados": ok}

@app.post("/import/persons/", tags=["Import/Export"])
def import_persons(file: UploadFile = File(...), db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_import_export"))):
    wb = openpyxl.load_workbook(file.file)
    ws = wb.active
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for col in ["Nombre", "Email", "EmployeeID"]:
        if col not in h: raise HTTPException(400, f"Falta columna '{col}'")
    def _ci(name): return h.index(name) if name in h else None
    i_fn, i_em, i_eid = h.index("Nombre"), h.index("Email"), h.index("EmployeeID")
    i_tit, i_tel, i_not, i_dept, i_sit = _ci("Titulo"), _ci("Telefono"), _ci("Notas"), _ci("Departamento"), _ci("Sitio")
    ok = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row): continue
        full_name = _cell(row[i_fn])
        email = _cell(row[i_em])
        eid = _cell(row[i_eid])
        if not full_name: raise HTTPException(400, f"Fila {ok+2}: Nombre completo vacio")
        if not email: raise HTTPException(400, f"Fila {ok+2}: Email vacio")
        if not eid: raise HTTPException(400, f"Fila {ok+2}: EmployeeID vacio")
        if db.query(models.Person).filter(models.Person.email == email).first(): raise HTTPException(400, f"Fila {ok+2}: Email '{email}' ya existe")
        if db.query(models.Person).filter(models.Person.employee_id == eid).first(): raise HTTPException(400, f"Fila {ok+2}: EmployeeID '{eid}' ya existe")
        dept_id, site_id = None, None
        if i_dept is not None and _cell(row[i_dept]):
            dept_name = _cell(row[i_dept])
            d = db.query(models.Department).filter(models.Department.department_name == dept_name).first()
            if not d:
                d = models.Department(department_name=dept_name)
                db.add(d); db.flush()
            dept_id = d.id
        if i_sit is not None and _cell(row[i_sit]):
            sit_name = _cell(row[i_sit])
            s = db.query(models.Site).filter(models.Site.site_name == sit_name).first()
            if not s:
                s = models.Site(site_name=sit_name)
                db.add(s); db.flush()
            site_id = s.id
        if dept_id is None: raise HTTPException(400, f"Fila {ok+2}: Departamento requerido")
        if site_id is None: raise HTTPException(400, f"Fila {ok+2}: Sitio requerido")
        db.add(models.Person(full_name=full_name, email=email, employee_id=eid, title=_cell(row[i_tit]) if i_tit is not None else None, phone=_cell(row[i_tel]) if i_tel is not None else None, notes=_cell(row[i_not]) if i_not is not None else None, department_id=dept_id, site_id=site_id))
        ok += 1
    db.commit()
    return {"importados": ok}

@app.post("/employees/reconcile/", tags=["Directorio"])
def reconcile_employees(file: UploadFile = File(...), db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_import_export"))):
    wb = openpyxl.load_workbook(file.file)
    ws = wb.active
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for col in ["Nombre", "Email", "EmployeeID"]:
        if col not in h: raise HTTPException(400, f"Falta columna '{col}'")
    def _ci(name): return h.index(name) if name in h else None
    i_fn, i_em, i_eid = h.index("Nombre"), h.index("Email"), h.index("EmployeeID")
    i_tit, i_tel, i_not, i_dept, i_sit = _ci("Titulo"), _ci("Telefono"), _ci("Notas"), _ci("Departamento"), _ci("Sitio")

    file_eids = set()
    file_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row): continue
        eid = _cell(row[i_eid])
        if not eid: continue
        file_eids.add(eid)
        file_rows.append({
            "full_name": _cell(row[i_fn]),
            "email": _cell(row[i_em]),
            "employee_id": eid,
            "title": _cell(row[i_tit]) if i_tit is not None else None,
            "phone": _cell(row[i_tel]) if i_tel is not None else None,
            "notes": _cell(row[i_not]) if i_not is not None else None,
            "dept_name": _cell(row[i_dept]) if i_dept is not None else None,
            "site_name": _cell(row[i_sit]) if i_sit is not None else None,
        })

    all_db_persons = db.query(models.Person).all()
    db_by_eid = {p.employee_id: p for p in all_db_persons if p.employee_id}

    departed_db_ids = set(db_by_eid.keys()) - file_eids
    new_file_ids = file_eids - set(db_by_eid.keys())

    # Auto-import new employees
    imported = []
    for row_data in file_rows:
        if row_data["employee_id"] not in new_file_ids: continue
        if not row_data["full_name"] or not row_data["email"]: continue
        dept_id, site_id = None, None
        if row_data["dept_name"]:
            d = db.query(models.Department).filter(models.Department.department_name == row_data["dept_name"]).first()
            if not d:
                d = models.Department(department_name=row_data["dept_name"])
                db.add(d); db.flush()
            dept_id = d.id
        if row_data["site_name"]:
            s = db.query(models.Site).filter(models.Site.site_name == row_data["site_name"]).first()
            if not s:
                s = models.Site(site_name=row_data["site_name"])
                db.add(s); db.flush()
            site_id = s.id
        new_p = models.Person(
            full_name=row_data["full_name"],
            email=row_data["email"],
            employee_id=row_data["employee_id"],
            title=row_data["title"],
            phone=row_data["phone"],
            notes=row_data["notes"],
            department_id=dept_id,
            site_id=site_id
        )
        db.add(new_p)
        db.flush()
        imported.append({"id": new_p.id, "full_name": new_p.full_name, "email": new_p.email, "employee_id": new_p.employee_id})
    db.commit()

    # Create reconciliation session
    session_rec = models.ReconciliationSession(
        uploaded_by_id=admin.id,
        filename=file.filename or "desconocido.xlsx",
        total_db=len(all_db_persons),
        total_file=len(file_eids),
        matched_count=len(file_eids & set(db_by_eid.keys())),
        imported_count=len(imported)
    )
    db.add(session_rec)
    db.flush()

    # Build departed list with only Checkout assets, save to BD
    departed = []
    departed_asset_records = []
    for eid in sorted(departed_db_ids):
        p = db_by_eid[eid]
        checkout_assets = db.query(models.Asset).filter(models.Asset.person_id == p.id, models.Asset.status == "Checkout").all()
        for a in checkout_assets:
            departed_asset_records.append(models.ReconciliationDepartedAsset(
                session_id=session_rec.id,
                person_id=p.id,
                asset_id=a.id,
                status="pending"
            ))
        if checkout_assets:
            departed.append({
                "person": {
                    "id": p.id, "full_name": p.full_name, "email": p.email,
                    "employee_id": p.employee_id, "title": p.title, "phone": p.phone
                },
                "assets": [
                    {
                        "id": a.id, "asset_tag_id": a.asset_tag_id,
                        "asset_description": a.asset_description, "brand": a.brand,
                        "model": a.model, "serial_no": a.serial_no,
                        "status": a.status, "category": a.category
                    }
                    for a in checkout_assets
                ]
            })
    for rec in departed_asset_records:
        db.add(rec)
    db.commit()

    return {
        "session_id": session_rec.id,
        "departed": departed,
        "new_employees": imported,
        "matched_count": session_rec.matched_count,
        "total_db": session_rec.total_db,
        "total_file": session_rec.total_file,
        "imported_count": session_rec.imported_count
    }

@app.get("/employees/reconciliation/status/", tags=["Directorio"])
def reconciliation_status(include_cleared: bool = False, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_import_export"))):
    sessions = db.query(models.ReconciliationSession).order_by(models.ReconciliationSession.uploaded_at.desc()).all()
    status_filter = [models.ReconciliationDepartedAsset.status == "pending"]
    if include_cleared:
        status_filter = []
    total_pending = db.query(models.ReconciliationDepartedAsset).filter(models.ReconciliationDepartedAsset.status == "pending").count()
    total_cleared = db.query(models.ReconciliationDepartedAsset).filter(models.ReconciliationDepartedAsset.status == "cleared").count()

    result = []
    for s in sessions:
        query = db.query(models.ReconciliationDepartedAsset).filter(models.ReconciliationDepartedAsset.session_id == s.id)
        if not include_cleared:
            query = query.filter(models.ReconciliationDepartedAsset.status == "pending")
        records = query.all()
        session_departed = {}
        for rec in records:
            p = rec.person
            a = rec.asset
            if p.id not in session_departed:
                session_departed[p.id] = {
                    "person": {"id": p.id, "full_name": p.full_name, "email": p.email, "employee_id": p.employee_id, "title": p.title},
                    "assets": []
                }
            session_departed[p.id]["assets"].append({
                "departed_asset_id": rec.id,
                "id": a.id, "asset_tag_id": a.asset_tag_id,
                "asset_description": a.asset_description, "brand": a.brand,
                "model": a.model, "serial_no": a.serial_no,
                "status": a.status, "category": a.category,
                "reconciliation_status": rec.status
            })
        result.append({
            "session_id": s.id,
            "uploaded_by": s.uploaded_by.username if s.uploaded_by else "Desconocido",
            "uploaded_at": s.uploaded_at.isoformat() if s.uploaded_at else None,
            "filename": s.filename,
            "total_db": s.total_db,
            "total_file": s.total_file,
            "matched_count": s.matched_count,
            "imported_count": s.imported_count,
            "departed": list(session_departed.values()),
            "pending_count": sum(1 for r in records if r.status == "pending"),
            "cleared_count": sum(1 for r in records if r.status == "cleared")
        })

    return {"sessions": result, "total_pending": total_pending, "total_cleared": total_cleared, "total_sessions": len(sessions)}

@app.post("/employees/reconciliation/{departed_asset_id}/clear/", tags=["Directorio"])
def reconciliation_clear(departed_asset_id: int, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_import_export"))):
    rec = db.query(models.ReconciliationDepartedAsset).filter(models.ReconciliationDepartedAsset.id == departed_asset_id).first()
    if not rec:
        raise HTTPException(404, "Registro de conciliacion no encontrado")
    rec.status = "cleared"
    rec.cleared_at = datetime.utcnow()
    rec.cleared_by_id = admin.id
    db.commit()
    return {"message": "Registro marcado como completado"}

@app.post("/employees/reconciliation/refresh/", tags=["Directorio"])
def reconciliation_refresh(db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_import_export"))):
    cleared = db.query(models.ReconciliationDepartedAsset).filter(
        models.ReconciliationDepartedAsset.status == "cleared"
    ).all()
    reactivated = 0
    for rec in cleared:
        asset = db.query(models.Asset).filter(models.Asset.id == rec.asset_id).first()
        if asset and asset.status == "Checkout" and asset.person_id == rec.person_id:
            rec.status = "pending"
            rec.cleared_at = None
            rec.cleared_by_id = None
            reactivated += 1
    db.commit()
    return {"reactivated": reactivated}

@app.post("/import/assets/", tags=["Import/Export"])
def import_assets(file: UploadFile = File(...), db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_import_export"))):
    wb = openpyxl.load_workbook(file.file)
    ws = wb.active
    h = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for col in ["AssetTag", "Descripcion", "Marca", "Modelo", "Serie", "Categoria"]:
        if col not in h: raise HTTPException(400, f"Falta columna '{col}'")
    i_tag, i_desc, i_brand, i_model, i_ser, i_cat = (h.index(c) for c in ["AssetTag", "Descripcion", "Marca", "Modelo", "Serie", "Categoria"])
    i_sit = h.index("Sitio") if "Sitio" in h else None
    i_asignado = h.index("AsignadoA") if "AsignadoA" in h else None
    ok = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row): continue
        tag = _cell(row[i_tag])
        desc = _cell(row[i_desc])
        brand = _cell(row[i_brand])
        model = _cell(row[i_model])
        ser = _cell(row[i_ser])
        cat_name = _cell(row[i_cat])
        sit_name = _cell(row[i_sit]) if i_sit is not None else ""
        if not tag: raise HTTPException(400, f"Fila {ok+2}: AssetTag vacio")
        if not desc: raise HTTPException(400, f"Fila {ok+2}: Descripcion vacia")
        if not brand: raise HTTPException(400, f"Fila {ok+2}: Marca vacia")
        if not model: raise HTTPException(400, f"Fila {ok+2}: Modelo vacio")
        if not ser: raise HTTPException(400, f"Fila {ok+2}: Serie vacia")
        if not cat_name: raise HTTPException(400, f"Fila {ok+2}: Categoria vacia")
        if sit_name and not db.query(models.Site).filter(models.Site.site_name == sit_name).first():
            raise HTTPException(400, f"Fila {ok+2}: Sitio '{sit_name}' no existe")
        if db.query(models.Asset).filter(models.Asset.asset_tag_id == tag).first(): raise HTTPException(400, f"Fila {ok+2}: AssetTag '{tag}' ya existe")
        site_id = db.query(models.Site.id).filter(models.Site.site_name == sit_name).scalar() if sit_name else None

        person_id = None
        status = "Available"
        notas_historial = None
        if i_asignado is not None:
            asignado = _cell(row[i_asignado])
            if asignado:
                person = db.query(models.Person).filter(models.Person.email == asignado).first()
                if not person:
                    raise HTTPException(400, f"Fila {ok+2}: Empleado con email '{asignado}' no encontrado")
                person_id = person.id
                status = "Checkout"
                notas_historial = f"Importado con asignacion a {person.full_name}"

        asset = models.Asset(asset_tag_id=tag, asset_description=desc, brand=brand, model=model, serial_no=ser, category=cat_name, site_id=site_id, person_id=person_id, status=status)
        db.add(asset)
        db.flush()

        if notas_historial:
            db.add(models.History(
                asset_id=asset.id, asignado_a_id=person_id, realizado_por_id=admin.id,
                tipo_accion="Checkout", estado_anterior="Available", estado_nuevo="Checkout",
                notas_detalle=notas_historial
            ))
        ok += 1
    db.commit()
    return {"importados": ok}

# ==========================================
# 12. ENDPOINTS DE REPORTES
# ==========================================
@app.get("/reports/person-checkouts/{person_id}", tags=["Reportes"])
def report_person_checkouts(person_id: int, mode: str = "current", db: Session = Depends(get_db)):
    person = db.query(models.Person).filter(models.Person.id == person_id).first()
    if not person:
        raise HTTPException(404, "Empleado no encontrado")

    if mode == "current":
        assets = db.query(models.Asset).filter(
            models.Asset.person_id == person_id,
            models.Asset.status == "Checkout"
        ).all()
        results = []
        for a in assets:
            site = db.query(models.Site).filter(models.Site.id == a.site_id).first()
            results.append(schemas.PersonCheckoutReportItem(
                asset_id=a.id,
                asset_tag_id=a.asset_tag_id,
                asset_description=a.asset_description,
                brand=a.brand,
                model=a.model,
                serial_no=a.serial_no,
                category=a.category or "",
                site_name=site.site_name if site else "",
                assigned_date=None,
                returned_date=None,
                status=a.status
            ))
        return results
    else:
        history = db.query(models.History).filter(
            models.History.asignado_a_id == person_id,
            models.History.tipo_accion.in_(["Checkout", "Check in"])
        ).order_by(models.History.asset_id, models.History.fecha_accion).all()
        asset_map = {}
        for h in history:
            aid = h.asset_id
            if aid not in asset_map:
                asset_map[aid] = {"checkout": None, "checkin": None}
            if h.tipo_accion == "Checkout" and asset_map[aid]["checkout"] is None:
                asset_map[aid]["checkout"] = h.fecha_accion
            if h.tipo_accion == "Check in":
                asset_map[aid]["checkin"] = h.fecha_accion
        asset_ids = list(asset_map.keys())
        assets = db.query(models.Asset).filter(models.Asset.id.in_(asset_ids)).all()
        asset_dict = {a.id: a for a in assets}
        results = []
        for aid, dates in asset_map.items():
            a = asset_dict.get(aid)
            if not a:
                continue
            site = db.query(models.Site).filter(models.Site.id == a.site_id).first()
            results.append(schemas.PersonCheckoutReportItem(
                asset_id=a.id,
                asset_tag_id=a.asset_tag_id,
                asset_description=a.asset_description,
                brand=a.brand,
                model=a.model,
                serial_no=a.serial_no,
                category=a.category or "",
                site_name=site.site_name if site else "",
                assigned_date=dates["checkout"].isoformat() if dates["checkout"] else None,
                returned_date=dates["checkin"].isoformat() if dates["checkin"] else None,
                status=a.status
            ))
        return results

@app.get("/reports/checkout-timeframe/", tags=["Reportes"])
def report_checkout_timeframe(
    start: str,
    end: str,
    db: Session = Depends(get_db)
):
    from datetime import datetime as dt_lib
    try:
        start_dt = dt_lib.strptime(start, "%Y-%m-%d")
        end_dt = dt_lib.strptime(end, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
    except ValueError:
        raise HTTPException(400, "Formato de fecha invalido. Use YYYY-MM-DD")

    history = db.query(models.History).filter(
        models.History.tipo_accion == "Checkout",
        models.History.fecha_accion >= start_dt,
        models.History.fecha_accion <= end_dt
    ).order_by(models.History.fecha_accion.desc()).all()

    asset_ids = list(set(h.asset_id for h in history))
    assets = {a.id: a for a in db.query(models.Asset).filter(models.Asset.id.in_(asset_ids)).all()}
    person_ids = list(set(h.asignado_a_id for h in history if h.asignado_a_id))
    persons = {p.id: p for p in db.query(models.Person).filter(models.Person.id.in_(person_ids)).all()}
    admin_ids = list(set(h.realizado_por_id for h in history))
    admins = {a.id: a for a in db.query(models.Admin).filter(models.Admin.id.in_(admin_ids)).all()}

    results = []
    for h in history:
        asset = assets.get(h.asset_id)
        if not asset:
            continue
        person = persons.get(h.asignado_a_id) if h.asignado_a_id else None
        admin = admins.get(h.realizado_por_id)
        results.append(schemas.CheckoutTimeframeItem(
            asset_id=asset.id,
            asset_tag_id=asset.asset_tag_id,
            asset_description=asset.asset_description,
            brand=asset.brand,
            model=asset.model,
            serial_no=asset.serial_no,
            category=asset.category or "",
            employee_name=person.full_name if person else "N/A",
            employee_id=person.employee_id if person else "",
            admin_name=admin.username if admin else "Desconocido",
            checkout_date=h.fecha_accion,
            current_status=asset.status
        ))
    return results

# ==========================================
# 13. ENDPOINTS DE ENTREGAS PENDIENTES
# ==========================================
AVAILABLE_STATUSES = ("Available",)

@app.get("/deliveries/available-assets", response_model=List[schemas.AvailableAssetItem], tags=["Entregas Pendientes"])
def list_available_assets(db: Session = Depends(get_db)):
    assets = db.query(models.Asset).filter(
        models.Asset.status.in_(AVAILABLE_STATUSES),
        models.Asset.person_id == None
    ).all()
    return assets

@app.post("/deliveries/pending", response_model=schemas.PendingDeliveryResponse, status_code=201, tags=["Entregas Pendientes"])
def create_pending_delivery(delivery: schemas.PendingDeliveryCreate, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_create"))):
    person = db.query(models.Person).filter(models.Person.id == delivery.person_id).first()
    if not person:
        raise HTTPException(404, "Empleado no encontrado")
    if delivery.quantity < 1:
        raise HTTPException(400, "La cantidad debe ser al menos 1")
    db_delivery = models.PendingDelivery(
        person_id=delivery.person_id,
        category=delivery.category,
        quantity=delivery.quantity,
        notes=delivery.notes
    )
    db.add(db_delivery)
    db.commit()
    db.refresh(db_delivery)
    resp = schemas.PendingDeliveryResponse.model_validate(db_delivery)
    resp.person_name = person.full_name
    return resp

@app.get("/deliveries/pending/", response_model=List[schemas.PendingDeliveryResponse], tags=["Entregas Pendientes"], include_in_schema=False)
@app.get("/deliveries/pending", response_model=List[schemas.PendingDeliveryResponse], tags=["Entregas Pendientes"])
def list_pending_deliveries(status: str = None, person_id: int = None, db: Session = Depends(get_db)):
    query = db.query(models.PendingDelivery)
    if status:
        query = query.filter(models.PendingDelivery.status == status)
    if person_id:
        query = query.filter(models.PendingDelivery.person_id == person_id)
    results = query.order_by(models.PendingDelivery.created_at.desc()).all()
    persons = {p.id: p.full_name for p in db.query(models.Person).all()}
    out = []
    for d in results:
        item = schemas.PendingDeliveryResponse.model_validate(d)
        item.person_name = persons.get(d.person_id, "")
        out.append(item)
    return out

@app.delete("/deliveries/pending/{delivery_id}", tags=["Entregas Pendientes"])
def cancel_pending_delivery(delivery_id: int, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_delete"))):
    d = db.query(models.PendingDelivery).filter(models.PendingDelivery.id == delivery_id).first()
    if not d:
        raise HTTPException(404, "Entrega pendiente no encontrada")
    d.status = "Cancelled"
    db.commit()
    return {"message": "Entrega pendiente cancelada"}

@app.post("/deliveries/pending/{delivery_id}/fulfill", tags=["Entregas Pendientes"])
def fulfill_pending_delivery(delivery_id: int, body: schemas.PendingFulfillRequest, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_checkout"))):
    d = db.query(models.PendingDelivery).filter(models.PendingDelivery.id == delivery_id).first()
    if not d:
        raise HTTPException(404, "Entrega pendiente no encontrada")
    if d.status != "Active":
        raise HTTPException(400, "Esta entrega ya no esta activa")
    if d.fulfilled_count >= d.quantity:
        raise HTTPException(400, "Todos los items ya fueron asignados")

    asset = db.query(models.Asset).filter(models.Asset.id == body.asset_id).first()
    if not asset:
        raise HTTPException(404, "Activo no encontrado")
    if asset.status not in AVAILABLE_STATUSES or asset.person_id is not None:
        raise HTTPException(400, "El activo no esta disponible para asignacion")
    if asset.category != d.category:
        raise HTTPException(400, f"El activo no pertenece a la categoria '{d.category}'")

    estado_anterior = asset.status
    asset.status = "Checkout"
    asset.person_id = d.person_id

    person = db.query(models.Person).filter(models.Person.id == d.person_id).first()
    person_name = person.full_name if person else ""
    db.add(models.History(
        asset_id=asset.id, asignado_a_id=d.person_id, realizado_por_id=admin.id,
        tipo_accion="Checkout", estado_anterior=estado_anterior, estado_nuevo="Checkout",
        notas_detalle=f"Entrega pendiente #{delivery_id}: {d.category} - {d.notes or person_name}"
    ))

    d.fulfilled_count += 1
    if d.fulfilled_count >= d.quantity:
        d.status = "Fulfilled"
        d.fulfilled_at = datetime.utcnow()

    db.commit()
    db.refresh(asset)
    return {"message": f"Asset {asset.asset_tag_id} asignado a la entrega pendiente #{delivery_id}", "asset_tag": asset.asset_tag_id, "fulfilled_count": d.fulfilled_count, "total": d.quantity}

@app.get("/deliveries/summary", tags=["Entregas Pendientes"])
def delivery_summary(db: Session = Depends(get_db)):
    results = db.query(models.PendingDelivery).filter(models.PendingDelivery.status == "Active").all()
    pending_assets = db.query(models.Asset).filter(
        models.Asset.status.in_(AVAILABLE_STATUSES),
        models.Asset.person_id == None
    ).all()
    available_map = {}
    for a in pending_assets:
        available_map[a.category] = available_map.get(a.category, 0) + 1
    categories = {}
    for d in results:
        cat = d.category
        if cat not in categories:
            categories[cat] = {"category": cat, "total_pending": 0, "available": 0, "employees": []}
        categories[cat]["total_pending"] += (d.quantity - d.fulfilled_count)
        categories[cat]["available"] = available_map.get(cat, 0)
        p = db.query(models.Person).filter(models.Person.id == d.person_id).first()
        categories[cat]["employees"].append({
            "delivery_id": d.id,
            "person_id": d.person_id,
            "person_name": p.full_name if p else "?",
            "pending": d.quantity - d.fulfilled_count,
            "notes": d.notes or ""
        })
    return list(categories.values())

# ==========================================
# 14. ENDPOINT DE AUTENTICACIÓN (LOGIN CON JWT)
# ==========================================
@app.post("/auth/login", response_model=schemas.TokenResponse, tags=["Autenticación"])
def login_admin(credentials: schemas.LoginRequest, db: Session = Depends(get_db)):
    admin = db.query(models.Admin).filter(models.Admin.username == credentials.username).first()

    if not admin or not verify_password(credentials.password, admin.password_hash):
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    
    if not admin.is_active:
        raise HTTPException(status_code=401, detail="Cuenta desactivada")

    token = create_access_token({"id": admin.id, "username": admin.username})
    
    group_resp = None
    if admin.group:
        group_resp = schemas.GroupResponse.model_validate(admin.group)
    
    return schemas.TokenResponse(
        access_token=token,
        admin=schemas.AdminResponse(
            id=admin.id, username=admin.username, email=admin.email,
            role=admin.role, group_id=admin.group_id, is_active=admin.is_active,
            group=group_resp
        )
    )

# ==========================================
# 15. ENDPOINTS DE GRUPOS Y USUARIOS
# ==========================================
@app.get("/groups/", response_model=List[schemas.GroupResponse], tags=["Grupos y Usuarios"])
def list_groups(db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_manage_users"))):
    return db.query(models.Group).all()

@app.post("/groups/", response_model=schemas.GroupResponse, status_code=201, tags=["Grupos y Usuarios"])
def create_group(group: schemas.GroupCreate, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_manage_users"))):
    existing = db.query(models.Group).filter(models.Group.name == group.name).first()
    if existing:
        raise HTTPException(400, "El grupo ya existe")
    db_group = models.Group(**group.model_dump())
    db.add(db_group)
    db.commit()
    db.refresh(db_group)
    return db_group

@app.put("/groups/{group_id}", response_model=schemas.GroupResponse, tags=["Grupos y Usuarios"])
def update_group(group_id: int, group: schemas.GroupUpdate, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_manage_users"))):
    db_group = db.query(models.Group).filter(models.Group.id == group_id).first()
    if not db_group:
        raise HTTPException(404, "Grupo no encontrado")
    for k, v in group.model_dump(exclude_unset=True).items():
        setattr(db_group, k, v)
    db.commit()
    db.refresh(db_group)
    return db_group

@app.delete("/groups/{group_id}", tags=["Grupos y Usuarios"])
def delete_group(group_id: int, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_manage_users"))):
    db_group = db.query(models.Group).filter(models.Group.id == group_id).first()
    if not db_group:
        raise HTTPException(404, "Grupo no encontrado")
    if db_group.is_default:
        raise HTTPException(400, "No se puede eliminar un grupo por defecto")
    users_in_group = db.query(models.Admin).filter(models.Admin.group_id == group_id).count()
    if users_in_group > 0:
        raise HTTPException(400, f"Hay {users_in_group} usuario(s) en este grupo. Reasignelos antes de eliminar.")
    db.delete(db_group)
    db.commit()
    return {"message": "Grupo eliminado"}

@app.get("/users/", response_model=List[schemas.AdminResponse], tags=["Grupos y Usuarios"])
def list_users(db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_manage_users"))):
    return db.query(models.Admin).all()

@app.post("/users/", response_model=schemas.AdminResponse, status_code=201, tags=["Grupos y Usuarios"])
def create_user(user: schemas.AdminCreate, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_manage_users"))):
    existing = db.query(models.Admin).filter((models.Admin.username == user.username) | (models.Admin.email == user.email)).first()
    if existing:
        raise HTTPException(400, "El nombre de usuario o email ya existe")
    db_user = models.Admin(
        username=user.username,
        email=user.email,
        password_hash=hash_password(user.password),
        role=user.role or "User",
        group_id=user.group_id,
        is_active=user.is_active
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

@app.put("/users/{user_id}", response_model=schemas.AdminResponse, tags=["Grupos y Usuarios"])
def update_user(user_id: int, user: schemas.AdminUpdate, db: Session = Depends(get_db), admin: models.Admin = Depends(require_permission("can_manage_users"))):
    db_user = db.query(models.Admin).filter(models.Admin.id == user_id).first()
    if not db_user:
        raise HTTPException(404, "Usuario no encontrado")
    update_data = user.model_dump(exclude_unset=True)
    if "password" in update_data and update_data["password"]:
        update_data["password_hash"] = hash_password(update_data.pop("password"))
    elif "password" in update_data:
        update_data.pop("password")
    for k, v in update_data.items():
        setattr(db_user, k, v)
    db.commit()
    db.refresh(db_user)
    return db_user

@app.post("/auth/change-password", tags=["Autenticación"])
def change_password(body: schemas.ChangePasswordRequest, db: Session = Depends(get_db), admin: models.Admin = Depends(get_current_admin)):
    if not verify_password(body.current_password, admin.password_hash):
        raise HTTPException(400, "La contrasena actual no es correcta")
    if len(body.new_password) < 4:
        raise HTTPException(400, "La nueva contrasena debe tener al menos 4 caracteres")
    admin.password_hash = hash_password(body.new_password)
    db.commit()
    return {"message": "Contrasena actualizada exitosamente"}